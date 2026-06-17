#!/usr/bin/env python3
"""
Build a P'sCUBE pachinko EV analysis file for HYPER ARROW Mihara.

This is a standalone verification CLI, not an appc integration.

The script can consume:
  - the existing P'sCUBE summary CSV
  - an optional history cache JSON generated from rendered P'sCUBE pages

Without rendered history rows, normal-start estimation is intentionally marked
as low confidence instead of pretending the P'sCUBE total start count is normal
rotation.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


OUTPUT_COLUMNS = [
    "date",
    "storeName",
    "machine",
    "dai",
    "totalStarts",
    "normalStartsMin",
    "normalStartsMid",
    "normalStartsMax",
    "normalStartsRange",
    "normalStartsRangeRate",
    "supportStartsMin",
    "supportStartsMid",
    "supportStartsMax",
    "jackpot",
    "continuations",
    "initialHits",
    "knownRunThroughCount",
    "unknownRunThroughCount",
    "finalStarts",
    "maxPayout",
    "estimatedDiffBalls",
    "diffBallsConfidence",
    "rotationRateMin",
    "rotationRateMid",
    "rotationRateMax",
    "expectedValueMin",
    "expectedValueMid",
    "expectedValueMax",
    "expectedHourlyMin",
    "expectedHourlyMid",
    "expectedHourlyMax",
    "confidence",
    "confidenceReason",
    "fetchStatus",
    "rankingEligible",
    "rankByExpectedHourlyMin",
    "rankByExpectedHourlyMid",
    "rankByRotationRateMin",
    "rankByRotationRateMid",
    "rankByNormalStarts",
    "rankByEstimatedDiffBalls",
    "url",
    "notes",
]


@dataclass(frozen=True)
class MachineSpec:
    key: str
    appa_name: str
    hatsua: float
    heikin: float
    heiren: float
    total: float
    jikan: float
    total1r: float
    st: int
    time_short: int
    support_candidates: Tuple[int, ...]


DEFAULT_SPECS: Dict[str, MachineSpec] = {
    "eva15": MachineSpec(
        key="eva15",
        appa_name="エヴァ",
        hatsua=319.7,
        heikin=1167.0,
        heiren=4.06,
        total=78.675,
        jikan=220.0,
        total1r=9.436,
        st=163,
        time_short=100,
        support_candidates=(100, 163, 500),
    ),
    "eva17": MachineSpec(
        key="eva17",
        appa_name="エヴァ17",
        hatsua=349.9,
        heikin=964.0,
        heiren=5.38,
        total=65.005,
        jikan=220.0,
        total1r=9.433,
        st=157,
        time_short=100,
        support_candidates=(100, 157),
    ),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate P'sCUBE Mihara Eva pachinko analysis CSV/XLSX."
    )
    parser.add_argument(
        "--input",
        default="data/pscube/samples/arrow_mihara_eva_4p_20260616.csv",
        help="Input CSV exported from P'sCUBE summary pages.",
    )
    parser.add_argument(
        "--history-cache",
        default="data/pscube/samples/pscube_history_cache_20260616.json",
        help="Optional rendered history cache JSON.",
    )
    parser.add_argument(
        "--diff-cache",
        default="data/pscube/samples/pscube_diff_cache_20260616.json",
        help="Optional graph-diff cache JSON. Values here fill missing estimatedDiffBalls.",
    )
    parser.add_argument(
        "--appa-masters",
        default="appa_masters_response.json",
        help="Optional appa masters JSON for spec/kokan reuse.",
    )
    parser.add_argument(
        "--out-csv",
        default="data/pscube/samples/pscube_mihara_eva_analysis_20260616.csv",
        help="Output analysis CSV.",
    )
    parser.add_argument(
        "--out-xlsx",
        default="data/pscube/samples/pscube_mihara_eva_analysis_20260616.xlsx",
        help="Output analysis XLSX. Requires openpyxl.",
    )
    parser.add_argument(
        "--summary-json",
        default="data/pscube/samples/pscube_mihara_eva_analysis_20260616_summary.json",
        help="Output summary JSON.",
    )
    parser.add_argument("--store-name", default="ハイパーアロー美原")
    parser.add_argument("--kokan", type=float, default=28.0, help="Exchange balls per 100 yen.")
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Skip XLSX creation.",
    )
    return parser.parse_args()


def to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def round_or_blank(value: Optional[float], digits: int = 0) -> Any:
    if value is None or not math.isfinite(value):
        return ""
    if digits == 0:
        return int(round(value))
    return round(value, digits)


def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in OUTPUT_COLUMNS})


def load_json(path: Path) -> Any:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def load_history_cache(path: Path) -> Dict[str, Any]:
    raw = load_json(path)
    if not raw:
        return {}
    if isinstance(raw, dict) and "machines" in raw:
        return {str(item.get("dai", "")).zfill(4): item for item in raw["machines"]}
    if isinstance(raw, dict):
        return {str(k).zfill(4): v for k, v in raw.items()}
    return {}


def load_diff_cache(path: Path) -> Dict[str, Dict[str, Any]]:
    raw = load_json(path)
    if not raw:
        return {}
    items: Iterable[Any]
    if isinstance(raw, dict) and "machines" in raw:
        items = raw["machines"]
    elif isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        normalized = {}
        for key, value in raw.items():
            dai = str(key).zfill(4)
            if isinstance(value, dict):
                normalized[dai] = value
            else:
                normalized[dai] = {"estimatedDiffBalls": value}
        return normalized
    else:
        return {}

    out: Dict[str, Dict[str, Any]] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        dai = str(item.get("dai", "")).zfill(4)
        if dai.strip("0") == "":
            continue
        out[dai] = item
    return out


def detect_machine_key(machine: str) -> str:
    if "17" in machine or "はじまり" in machine:
        return "eva17"
    return "eva15"


def load_appa_specs(path: Path) -> Tuple[Dict[str, MachineSpec], Optional[float], str]:
    specs = dict(DEFAULT_SPECS)
    kokan = None
    note = "appa masters not found; used embedded fallback specs"
    raw = load_json(path)
    if not raw:
        return specs, kokan, note

    kishus = raw.get("kishus") or []
    by_name = {str(k.get("name")): k for k in kishus}

    def merged(key: str, appa_name: str) -> MachineSpec:
        base = specs[key]
        row = by_name.get(appa_name)
        if not row:
            return base
        return MachineSpec(
            key=base.key,
            appa_name=appa_name,
            hatsua=float(row.get("hatsua") or base.hatsua),
            heikin=float(row.get("heikin") or base.heikin),
            heiren=float(row.get("heiren") or base.heiren),
            total=float(row.get("total") or base.total),
            jikan=float(row.get("jikan") or base.jikan),
            total1r=float(row.get("total1R") or row.get("total1r") or base.total1r),
            st=base.st,
            time_short=base.time_short,
            support_candidates=base.support_candidates,
        )

    specs["eva15"] = merged("eva15", "エヴァ")
    specs["eva17"] = merged("eva17", "エヴァ17")

    for shop in raw.get("shops") or []:
        if shop.get("name") == "美原":
            try:
                kokan = float(shop.get("kokan"))
            except (TypeError, ValueError):
                kokan = None
            break

    note = "appa masters loaded; reused エヴァ / エヴァ17 hatsua, heikin, heiren, total, jikan, total1R"
    return specs, kokan, note


def normalize_history_rows(cache_item: Optional[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], str]:
    if not cache_item:
        return [], "history_cache_missing"
    status = cache_item.get("fetchStatus") or cache_item.get("status") or "history_cache_loaded"
    if status not in ("ok", "skipped_no_jackpot"):
        return [], str(status)

    rows = cache_item.get("historyRows") or cache_item.get("rows") or []
    normalized = []
    for row in rows:
        if isinstance(row, dict):
            normalized.append(
                {
                    "bonusId": to_int(row.get("bonusId")),
                    "time": row.get("time") or row.get("hhmm") or "",
                    "game": to_int(row.get("game")) or 0,
                    "payout": row.get("payout") or row.get("sum_dedama") or "",
                    "status": str(row.get("status") or row.get("nmk_status") or ""),
                }
            )
        elif isinstance(row, list) and len(row) >= 5:
            normalized.append(
                {
                    "bonusId": to_int(row[0]),
                    "time": row[1],
                    "game": to_int(row[2]) or 0,
                    "payout": row[3],
                    "status": str(row[4]),
                }
            )
    normalized = [r for r in normalized if r["bonusId"] is not None]
    normalized.sort(key=lambda r: r["bonusId"])
    if status == "ok" and not normalized:
        return [], "history_empty_for_target_date"
    return normalized, str(status)


def split_chains(history_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    chains: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None
    for row in history_rows:
        status = row.get("status", "")
        if "初" in status:
            if current:
                chains.append(current)
            current = {"initial": row, "continuations": []}
        else:
            if current is None:
                current = {"initial": None, "continuations": []}
            current["continuations"].append(row)
    if current:
        chains.append(current)
    return chains


def support_candidates_for_last_chain(
    spec: MachineSpec, chains: List[Dict[str, Any]], final_starts: int, jackpot: int
) -> List[int]:
    if jackpot <= 0 or final_starts <= 0:
        return [0]
    if not chains:
        return [0]
    last = chains[-1]
    cont_count = len(last.get("continuations") or [])
    if cont_count > 0:
        return [min(final_starts, spec.st)]
    return sorted({min(final_starts, spec.time_short), min(final_starts, spec.st)})


def estimate_normal_starts(
    row: Dict[str, str], spec: MachineSpec, history_rows: List[Dict[str, Any]], fetch_status: str
) -> Dict[str, Any]:
    total_starts = to_int(row.get("totalStarts")) or 0
    final_starts = to_int(row.get("finalStarts")) or 0
    jackpot = to_int(row.get("jackpot")) or 0

    if jackpot == 0:
        normal = total_starts
        return {
            "normal_min": normal,
            "normal_mid": normal,
            "normal_max": normal,
            "support_min": 0,
            "support_mid": 0,
            "support_max": 0,
            "known_runthrough": 0,
            "unknown_runthrough": 0,
            "confidence": "A",
            "reason": "大当たり0のため電サポ控除なし",
            "notes": "履歴不要で通常回転は累計スタート相当と推定",
        }

    if fetch_status != "ok" or not history_rows:
        initial_hits = to_int(row.get("initialHits")) or 0
        continuations = to_int(row.get("continuations")) or 0

        # Summary-only fallback. This is intentionally a wide range: it keeps
        # diff-ball review usable without pretending the P'sCUBE total start
        # count is confirmed normal rotation.
        support_unit_mid = (spec.time_short + spec.st) / 4
        continuation_mid = spec.st / 2
        support_mid = min(
            float(total_starts),
            initial_hits * support_unit_mid + continuations * continuation_mid,
        )
        support_band = max(float(spec.st), support_mid * 0.35)
        support_min = max(0.0, support_mid - support_band)
        support_max = min(float(total_starts), support_mid + support_band)
        normal_min = max(0.0, float(total_starts) - support_max)
        normal_max = max(normal_min, float(total_starts) - support_min)
        normal_mid = (normal_min + normal_max) / 2
        normal_range = normal_max - normal_min
        range_rate = normal_range / normal_mid if normal_mid else 1.0
        confidence = "C-" if range_rate <= 0.20 else "D"
        return {
            "normal_min": normal_min,
            "normal_mid": normal_mid,
            "normal_max": normal_max,
            "support_min": support_min,
            "support_mid": support_mid,
            "support_max": support_max,
            "known_runthrough": 0,
            "unknown_runthrough": max(1, initial_hits),
            "confidence": confidence,
            "reason": (
                f"履歴なしのため台一覧サマリーから暫定レンジ化: {fetch_status}; "
                f"supportMid={support_mid:.1f}, supportBand={support_band:.1f}"
            ),
            "notes": "履歴なし暫定推定。差玉確認用で、回転率/期待値は要再検証",
        }

    chains = split_chains(history_rows)
    initial_game_sum = sum(r["game"] for r in history_rows if "初" in r.get("status", ""))
    continuation_game_sum = sum(r["game"] for r in history_rows if "継続" in r.get("status", ""))
    observed_sum = initial_game_sum + continuation_game_sum
    missing_support = total_starts - final_starts - observed_sum
    if missing_support < 0:
        missing_support = 0

    candidates = support_candidates_for_last_chain(spec, chains, final_starts, jackpot)
    final_normal_candidates = [max(0, final_starts - support) for support in candidates]

    normal_min = initial_game_sum + min(final_normal_candidates)
    normal_max = initial_game_sum + max(final_normal_candidates)
    normal_mid = (normal_min + normal_max) / 2
    support_min = total_starts - normal_max
    support_max = total_starts - normal_min
    support_mid = (support_min + support_max) / 2

    unknown = 1 if len(set(final_normal_candidates)) > 1 else 0
    known = 1 if missing_support > 0 else 0
    normal_range = normal_max - normal_min
    range_rate = normal_range / normal_mid if normal_mid else 1

    if range_rate <= 0.03 and unknown == 0:
        confidence = "A"
    elif range_rate <= 0.03:
        confidence = "B"
    elif range_rate <= 0.07:
        confidence = "C"
    elif range_rate <= 0.10:
        confidence = "C-"
    else:
        confidence = "D"

    reason = (
        f"履歴成功。初当りgame合計={initial_game_sum}, 継続game合計={continuation_game_sum}, "
        f"未表示電サポ推定={missing_support}, 最終候補={candidates}"
    )

    return {
        "normal_min": normal_min,
        "normal_mid": normal_mid,
        "normal_max": normal_max,
        "support_min": support_min,
        "support_mid": support_mid,
        "support_max": support_max,
        "known_runthrough": known,
        "unknown_runthrough": unknown,
        "confidence": confidence,
        "reason": reason,
        "notes": "履歴DOMから推定。最終スタート中の時短/ST区別が曖昧な場合はレンジ化",
    }


def estimate_rotation_and_ev(
    normal_starts: Optional[float],
    diff_balls: Optional[int],
    max_payout: int,
    initial_hits: int,
    spec: MachineSpec,
    kokan: float,
) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if normal_starts is None or normal_starts <= 0 or diff_balls is None:
        return None, None, None

    theoretical_gross = initial_hits * spec.heikin * spec.heiren
    gross_balls = max(float(max_payout or 0), theoretical_gross)
    used_balls = gross_balls - diff_balls
    if used_balls <= 0:
        return None, None, None

    rotation = normal_starts / (used_balls / 250.0)
    if rotation <= 0:
        return None, None, None

    kankin = 100.0 / kokan
    zenki = spec.heikin * spec.jikan / spec.total * kankin
    koki = 250.0 * spec.jikan / rotation * kankin
    expected_hourly = zenki - koki
    expected_value = expected_hourly * (normal_starts / spec.jikan)
    return rotation, expected_value, expected_hourly


def classify_diff_confidence(value: Optional[int]) -> str:
    return "graph_estimated" if value is not None else "missing"


def build_analysis_rows(
    input_rows: List[Dict[str, str]],
    history_cache: Dict[str, Any],
    diff_cache: Dict[str, Dict[str, Any]],
    specs: Dict[str, MachineSpec],
    store_name: str,
    kokan: float,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for src in input_rows:
        dai = str(src.get("dai", "")).zfill(4)
        spec = specs[detect_machine_key(src.get("machine", ""))]
        cache_item = history_cache.get(dai)
        hist_rows, fetch_status = normalize_history_rows(cache_item)
        total_starts = to_int(src.get("totalStarts")) or 0
        if fetch_status == "ok" and cache_item:
            # When the site has rolled to the next business day, the table can
            # show the target date in the "1日前" column while #tblHISTb still
            # contains only today's history. Reject that mixed-date history.
            current_total = None
            for cells in cache_item.get("dataRows") or []:
                if isinstance(cells, list) and cells and cells[0] == "累計スタート":
                    current_total = to_int(cells[1] if len(cells) > 1 else None)
                    break
            if current_total is not None and current_total != total_starts:
                hist_rows = []
                fetch_status = "history_not_available_for_past_date"
        estimation = estimate_normal_starts(src, spec, hist_rows, fetch_status)

        jackpot = to_int(src.get("jackpot")) or 0
        continuations = to_int(src.get("continuations")) or 0
        initial_hits = to_int(src.get("initialHits")) or 0
        final_starts = to_int(src.get("finalStarts")) or 0
        max_payout = to_int(src.get("maxPayout")) or 0
        diff_balls = to_int(src.get("estimatedDiffBalls"))
        diff_confidence = classify_diff_confidence(diff_balls)
        diff_item = diff_cache.get(dai, {})
        cached_diff = to_int(diff_item.get("estimatedDiffBalls") or diff_item.get("diffBalls"))
        if diff_balls is None and cached_diff is not None:
            diff_balls = cached_diff
            diff_confidence = str(diff_item.get("diffBallsConfidence") or "raw_graph_json")

        normal_min = estimation["normal_min"]
        normal_mid = estimation["normal_mid"]
        normal_max = estimation["normal_max"]
        normal_range = (
            None if normal_min is None or normal_max is None else float(normal_max) - float(normal_min)
        )
        normal_range_rate = (
            None
            if normal_range is None or not normal_mid
            else normal_range / float(normal_mid)
        )

        rot_min, ev_min, hourly_min = estimate_rotation_and_ev(
            normal_min, diff_balls, max_payout, initial_hits, spec, kokan
        )
        rot_mid, ev_mid, hourly_mid = estimate_rotation_and_ev(
            normal_mid, diff_balls, max_payout, initial_hits, spec, kokan
        )
        rot_max, ev_max, hourly_max = estimate_rotation_and_ev(
            normal_max, diff_balls, max_payout, initial_hits, spec, kokan
        )

        confidence = estimation["confidence"]
        eligible = confidence not in ("D",) and hourly_min is not None

        notes = [
            estimation["notes"],
            f"spec={spec.appa_name}",
            "rotation uses estimated used balls = max(maxPayout, initialHits*heikin*heiren) - estimatedDiffBalls",
        ]
        if diff_balls is None:
            notes.append("estimatedDiffBalls missing; rotation/EV blank")
        elif diff_item:
            notes.append(f"estimatedDiffBalls filled from diff cache ({diff_confidence})")

        out.append(
            {
                "date": src.get("date", ""),
                "storeName": store_name,
                "machine": src.get("machine", ""),
                "dai": dai,
                "totalStarts": total_starts,
                "normalStartsMin": round_or_blank(normal_min),
                "normalStartsMid": round_or_blank(normal_mid, 1),
                "normalStartsMax": round_or_blank(normal_max),
                "normalStartsRange": round_or_blank(normal_range),
                "normalStartsRangeRate": round_or_blank(normal_range_rate, 4),
                "supportStartsMin": round_or_blank(estimation["support_min"]),
                "supportStartsMid": round_or_blank(estimation["support_mid"], 1),
                "supportStartsMax": round_or_blank(estimation["support_max"]),
                "jackpot": jackpot,
                "continuations": continuations,
                "initialHits": initial_hits,
                "knownRunThroughCount": estimation["known_runthrough"],
                "unknownRunThroughCount": estimation["unknown_runthrough"],
                "finalStarts": final_starts,
                "maxPayout": max_payout,
                "estimatedDiffBalls": "" if diff_balls is None else diff_balls,
                "diffBallsConfidence": diff_confidence,
                "rotationRateMin": round_or_blank(rot_min, 2),
                "rotationRateMid": round_or_blank(rot_mid, 2),
                "rotationRateMax": round_or_blank(rot_max, 2),
                "expectedValueMin": round_or_blank(ev_min),
                "expectedValueMid": round_or_blank(ev_mid),
                "expectedValueMax": round_or_blank(ev_max),
                "expectedHourlyMin": round_or_blank(hourly_min),
                "expectedHourlyMid": round_or_blank(hourly_mid),
                "expectedHourlyMax": round_or_blank(hourly_max),
                "confidence": confidence,
                "confidenceReason": estimation["reason"],
                "fetchStatus": fetch_status,
                "rankingEligible": "TRUE" if eligible else "FALSE",
                "rankByExpectedHourlyMin": "",
                "rankByExpectedHourlyMid": "",
                "rankByRotationRateMin": "",
                "rankByRotationRateMid": "",
                "rankByNormalStarts": "",
                "rankByEstimatedDiffBalls": "",
                "url": src.get("url", ""),
                "notes": " | ".join(notes),
            }
        )
    add_ranks(out)
    return out


def sortable_number(row: Dict[str, Any], key: str) -> Optional[float]:
    value = row.get(key)
    if value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def add_rank(
    rows: List[Dict[str, Any]], metric: str, rank_col: str, require_eligible: bool = True
) -> None:
    ranked = [
        row
        for row in rows
        if (not require_eligible or row.get("rankingEligible") == "TRUE")
        and sortable_number(row, metric) is not None
    ]
    ranked.sort(key=lambda r: sortable_number(r, metric), reverse=True)
    for idx, row in enumerate(ranked, 1):
        row[rank_col] = idx


def add_ranks(rows: List[Dict[str, Any]]) -> None:
    add_rank(rows, "expectedHourlyMin", "rankByExpectedHourlyMin")
    add_rank(rows, "expectedHourlyMid", "rankByExpectedHourlyMid")
    add_rank(rows, "rotationRateMin", "rankByRotationRateMin")
    add_rank(rows, "rotationRateMid", "rankByRotationRateMid")
    add_rank(rows, "normalStartsMid", "rankByNormalStarts")
    add_rank(rows, "estimatedDiffBalls", "rankByEstimatedDiffBalls", require_eligible=False)


def average(values: Iterable[Optional[float]]) -> Optional[float]:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    if not clean:
        return None
    return sum(clean) / len(clean)


def build_summary(rows: List[Dict[str, Any]], appa_note: str, kokan: float) -> Dict[str, Any]:
    counts: Dict[str, int] = {}
    for row in rows:
        counts[row["confidence"]] = counts.get(row["confidence"], 0) + 1

    def top(metric: str, limit: int = 10) -> List[Dict[str, Any]]:
        ranked = [r for r in rows if sortable_number(r, metric) is not None]
        ranked.sort(key=lambda r: sortable_number(r, metric), reverse=True)
        return [
            {
                "rank": idx,
                "dai": r["dai"],
                "machine": r["machine"],
                metric: r[metric],
                "confidence": r["confidence"],
                "normalStartsMid": r["normalStartsMid"],
                "rotationRateMid": r["rotationRateMid"],
            }
            for idx, r in enumerate(ranked[:limit], 1)
        ]

    by_machine: Dict[str, Dict[str, Any]] = {}
    for machine in sorted({r["machine"] for r in rows}):
        group = [r for r in rows if r["machine"] == machine]
        by_machine[machine] = {
            "count": len(group),
            "avgRotationRateMid": round_or_blank(
                average([sortable_number(r, "rotationRateMid") for r in group]), 2
            ),
            "avgExpectedHourlyMid": round_or_blank(
                average([sortable_number(r, "expectedHourlyMid") for r in group]), 0
            ),
        }

    normal_min_sum = sum(sortable_number(r, "normalStartsMin") or 0 for r in rows)
    normal_mid_sum = sum(sortable_number(r, "normalStartsMid") or 0 for r in rows)
    normal_max_sum = sum(sortable_number(r, "normalStartsMax") or 0 for r in rows)

    return {
        "targetCount": len(rows),
        "historySuccessCount": sum(1 for r in rows if r["fetchStatus"] == "ok"),
        "historyFailureCount": sum(1 for r in rows if r["fetchStatus"] != "ok"),
        "confidenceCounts": counts,
        "rankingEligibleCount": sum(1 for r in rows if r["rankingEligible"] == "TRUE"),
        "estimatedDiffBallsCount": sum(1 for r in rows if r["estimatedDiffBalls"] != ""),
        "topExpectedHourlyMin": top("expectedHourlyMin"),
        "topExpectedHourlyMid": top("expectedHourlyMid"),
        "topEstimatedDiffBalls": top("estimatedDiffBalls"),
        "normalStartsTotal": {
            "min": round(normal_min_sum, 1),
            "mid": round(normal_mid_sum, 1),
            "max": round(normal_max_sum, 1),
        },
        "byMachine": by_machine,
        "appaFormulaReuse": {
            "status": "partial_reuse",
            "note": appa_note,
            "formula": "border=hatsua/(heikin*heiren/250); expectedHourly=(heikin*jikan/total - 250*jikan/rotationRate)*(100/kokan)",
            "limitation": "Appa's exact rotation formula needs used balls from investment/recovery. This batch file estimates used balls from maxPayout/theoretical gross and graph-estimated diff.",
        },
        "kokan": kokan,
    }


def write_xlsx(path: Path, rows: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(f"openpyxl is not available: {exc}") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    table = Table(displayName="PscubeAnalysis", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    widths = {
        "A": 12, "B": 18, "C": 28, "D": 8, "E": 12, "F": 15, "G": 15, "H": 15,
        "AI": 50, "AP": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col in range(1, ws.max_column + 1):
        letter = ws.cell(1, col).column_letter
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 14

    sm = wb.create_sheet("Summary")
    sm.append(["Metric", "Value"])
    for key in [
        "targetCount",
        "historySuccessCount",
        "historyFailureCount",
        "rankingEligibleCount",
        "estimatedDiffBallsCount",
        "kokan",
    ]:
        sm.append([key, summary.get(key)])
    sm.append([])
    sm.append(["Confidence", "Count"])
    for key, value in sorted((summary.get("confidenceCounts") or {}).items()):
        sm.append([key, value])
    sm.append([])
    sm.append(["Top expectedHourlyMin", ""])
    sm.append(["rank", "dai", "machine", "expectedHourlyMin", "confidence", "rotationRateMid"])
    for item in summary.get("topExpectedHourlyMin", []):
        sm.append([
            item["rank"], item["dai"], item["machine"],
            item["expectedHourlyMin"], item["confidence"], item["rotationRateMid"],
        ])
    for cell in sm[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sm.column_dimensions["A"].width = 24
    sm.column_dimensions["B"].width = 18
    sm.column_dimensions["C"].width = 28
    sm.freeze_panes = "A2"

    wb.save(path)


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    history_path = Path(args.history_cache)
    diff_path = Path(args.diff_cache)
    masters_path = Path(args.appa_masters)
    out_csv = Path(args.out_csv)
    out_xlsx = Path(args.out_xlsx)
    summary_json = Path(args.summary_json)

    input_rows = read_csv(input_path)
    history_cache = load_history_cache(history_path)
    diff_cache = load_diff_cache(diff_path)
    specs, masters_kokan, appa_note = load_appa_specs(masters_path)
    kokan = masters_kokan or args.kokan

    rows = build_analysis_rows(input_rows, history_cache, diff_cache, specs, args.store_name, kokan)
    summary = build_summary(rows, appa_note, kokan)

    write_csv(out_csv, rows)
    with summary_json.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    if not args.no_xlsx:
        write_xlsx(out_xlsx, rows, summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

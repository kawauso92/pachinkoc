#!/usr/bin/env python3
"""
P'sCUBE diff-ball investigation script using Playwright.

Investigates all available routes to extract diff-ball (差玉) data from P'sCUBE
machine pages for HYPER ARROW Mihara.

Routes investigated:
  1. nc-m06-001.php JSON (main API) - Graph[].src.datas last value
  2. nc-m06-003.php JSON (AmChart7 API) - Graph.src.datas.p last point
  3. SVG/Canvas DOM extraction after rendering
  4. Graph image URL availability
  5. Past-date parameter variations

Safe operation:
  - Playwright browser with realistic user agent
  - 8-second delay between machines
  - Stops on 429/451/1015
  - Saves all responses to debug_cache/
  - Reuses cached responses on --resume
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

BLOCK_PATTERNS = ["429", "451", "1015", "captcha", "challenge-platform"]
STORE_ID = "c732925"
BASE_PATH = f"/dedamajyoho-P-townDMMpachi/{STORE_ID}/cgi-bin"


@dataclass
class MachineResult:
    date: str = ""
    machine: str = ""
    dai: str = ""
    url: str = ""
    totalStarts: int = 0
    jackpot: int = 0
    continuations: int = 0
    initialHits: int = 0
    maxPayout: int = 0
    existingEstimatedDiffBalls: str = ""
    diffBallsFromApi: str = ""
    diffBallsFromSvg: str = ""
    diffBallsFinal: str = ""
    diffBallsMethod: str = ""
    diffBallsConfidence: str = ""
    graphEndpointUsed: str = ""
    graphPointCount: int = 0
    graphLastPointRaw: str = ""
    graphZeroLine: str = ""
    graphScaleInfo: str = ""
    fetchStatus: str = ""
    httpStatus: str = ""
    errorType: str = ""
    errorMessage: str = ""
    notes: str = ""
    apiResponseKeys: str = ""
    graphType: str = ""
    ymdBizEmbedded: str = ""
    m06_001_status: str = ""
    m06_003_status: str = ""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="P'sCUBE diff-ball investigation")
    p.add_argument("--input", default="data/pscube/samples/arrow_mihara_eva_4p_20260616.csv")
    p.add_argument("--date", default="20260616")
    p.add_argument("--out-csv", default="data/pscube/samples/pscube_mihara_eva_diff_debug_20260616.csv")
    p.add_argument("--out-xlsx", default="data/pscube/samples/pscube_mihara_eva_diff_debug_20260616.xlsx")
    p.add_argument("--cache-dir", default="debug_cache")
    p.add_argument("--delay", type=float, default=8.0)
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--resume", action="store_true")
    p.add_argument("--headless", action="store_true", default=True)
    p.add_argument("--no-headless", dest="headless", action="store_false")
    p.add_argument("--single", type=str, default="", help="Test single dai")
    p.add_argument("--skip-browser", action="store_true",
                   help="Only analyze cached data, no browser fetch")
    return p.parse_args()


def to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def read_csv_input(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_json(path: Path) -> Any:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def extract_tokens_from_html(html: str) -> Optional[Dict[str, str]]:
    apikey_m = re.search(r"api\.apikey\s*=\s*'([^']+)'", html)
    token_m = re.search(r"api\.token\s*=\s*JSON\.parse\('([^']+)'\)", html)
    if not apikey_m or not token_m:
        return None
    try:
        tok = json.loads(token_m.group(1))
    except json.JSONDecodeError:
        return None
    _i = str(tok.get("_i", ""))
    _t = str(tok.get("_t", ""))
    if not _i or not _t:
        return None
    return {"apikey": apikey_m.group(1), "_i": _i, "_t": _t}


def extract_ymd_biz_from_html(html: str) -> str:
    m = re.search(r"api06\s*\.\s*show\s*\(\s*'[^']*'\s*,\s*'(\d{8})'", html)
    return m.group(1) if m else ""


def extract_model_from_html(html: str) -> str:
    m = re.search(r"api\.model\s*=\s*'([^']+)'", html)
    return m.group(1) if m else "nc-m06-001.php"


def extract_diff_from_m06_001(data: Any) -> Tuple[Optional[int], int, str, str, str]:
    """Extract diff balls from nc-m06-001.php response.
    Returns (diff_balls, point_count, last_point_raw, graph_type, scale_info)
    """
    if not isinstance(data, dict):
        return None, 0, "", "", ""

    graph_type = str(data.get("GraphType", ""))
    graphs = data.get("Graph")
    if graphs is None:
        return None, 0, "", graph_type, ""

    if isinstance(graphs, dict):
        graphs = [graphs]

    for graph in graphs:
        if not isinstance(graph, dict):
            continue
        src = graph.get("src")
        if not isinstance(src, dict):
            continue
        datas = src.get("datas")
        if datas is None:
            continue

        scale_info = ""
        y_axis = src.get("yAxis", {})
        if isinstance(y_axis, dict):
            y_min = y_axis.get("minimum")
            y_max = y_axis.get("maximum")
            scale_info = f"yMin={y_min},yMax={y_max}"

        if isinstance(datas, list) and len(datas) > 0:
            last = datas[-1]
            if isinstance(last, dict):
                val = last.get("value")
                if val is not None:
                    try:
                        return int(round(float(val))), len(datas), json.dumps(last, ensure_ascii=False), graph_type, scale_info
                    except (ValueError, TypeError):
                        pass
                raw = json.dumps(last, ensure_ascii=False)
                return None, len(datas), raw, graph_type, scale_info
            elif isinstance(last, (int, float)):
                return int(round(last)), len(datas), str(last), graph_type, scale_info

        elif isinstance(datas, dict):
            points = datas.get("p", [])
            g_defs = datas.get("g", [])
            y_fields = []
            for gd in g_defs:
                if isinstance(gd, dict) and gd.get("yField"):
                    y_fields.append(str(gd["yField"]))
            if not y_fields:
                y_fields = ["value", "y"]
            if isinstance(points, list) and len(points) > 0:
                last = points[-1]
                if isinstance(last, dict):
                    for yf in y_fields:
                        val = last.get(yf)
                        if val is not None:
                            try:
                                return int(round(float(val))), len(points), json.dumps(last, ensure_ascii=False), graph_type, scale_info
                            except (ValueError, TypeError):
                                pass
                    raw = json.dumps(last, ensure_ascii=False)
                    return None, len(points), raw, graph_type, scale_info

    return None, 0, "", graph_type, ""


def extract_diff_from_m06_003(data: Any) -> Tuple[Optional[int], int, str, str]:
    """Extract diff balls from nc-m06-003.php (AmChart7) response.
    Returns (diff_balls, point_count, last_point_raw, scale_info)
    """
    if not isinstance(data, dict):
        return None, 0, "", ""

    graph = data.get("Graph")
    if not isinstance(graph, dict):
        return None, 0, "", ""

    src = graph.get("src")
    if not isinstance(src, dict):
        return None, 0, "", ""

    scale_info = ""
    y_axis = src.get("yAxis", {})
    if isinstance(y_axis, dict):
        scale_info = f"yMin={y_axis.get('minimum')},yMax={y_axis.get('maximum')}"

    datas = src.get("datas")
    if not isinstance(datas, dict):
        return None, 0, "", scale_info

    points = datas.get("p", [])
    g_defs = datas.get("g", [])
    y_fields = []
    for gd in g_defs:
        if isinstance(gd, dict) and gd.get("yField"):
            y_fields.append(str(gd["yField"]))
    if not y_fields:
        y_fields = ["value", "y"]

    if not isinstance(points, list) or len(points) == 0:
        return None, 0, "", scale_info

    last = points[-1]
    if isinstance(last, dict):
        for yf in y_fields:
            val = last.get(yf)
            if val is not None:
                try:
                    return int(round(float(val))), len(points), json.dumps(last, ensure_ascii=False), scale_info
                except (ValueError, TypeError):
                    pass
        raw = json.dumps(last, ensure_ascii=False)
        return None, len(points), raw, scale_info

    return None, len(points), str(last), scale_info


def extract_svg_diff(page: Any) -> Tuple[Optional[int], str, str]:
    """Try to extract diff ball value from rendered SVG/Canvas on page."""
    try:
        result = page.evaluate("""() => {
            const info = {svgElements: 0, canvasElements: 0, amchartsFound: false,
                          lastValue: null, yAxisLabels: [], zeroLine: null};

            // Check for AmCharts instances
            if (typeof AmCharts !== 'undefined' && AmCharts.charts) {
                info.amchartsFound = true;
                for (const chart of AmCharts.charts) {
                    if (chart.dataProvider && chart.dataProvider.length > 0) {
                        const last = chart.dataProvider[chart.dataProvider.length - 1];
                        if (last.value !== undefined) {
                            info.lastValue = last.value;
                        }
                        // Check y-axis
                        if (chart.valueAxes) {
                            for (const axis of chart.valueAxes) {
                                if (axis.minimum !== undefined) {
                                    info.yAxisLabels.push({min: axis.minimum, max: axis.maximum});
                                }
                            }
                        }
                    }
                }
            }

            // Check SVG elements
            const svgs = document.querySelectorAll('svg');
            info.svgElements = svgs.length;
            for (const svg of svgs) {
                const texts = svg.querySelectorAll('text');
                for (const t of texts) {
                    const val = t.textContent.trim().replace(/,/g, '');
                    if (/^-?\d+$/.test(val) && Math.abs(parseInt(val)) > 100) {
                        info.yAxisLabels.push(val);
                    }
                }
            }

            // Check canvas elements
            info.canvasElements = document.querySelectorAll('canvas').length;

            return info;
        }""")

        if result and result.get("lastValue") is not None:
            val = result["lastValue"]
            try:
                return int(round(float(val))), "amcharts_dom", json.dumps(result, ensure_ascii=False)
            except (ValueError, TypeError):
                pass

        return None, "svg_extraction_failed", json.dumps(result or {}, ensure_ascii=False)
    except Exception as e:
        return None, "svg_error", str(e)


def investigate_machine_with_browser(
    page: Any,
    dai: str,
    target_ymd: str,
    cache_dir: Path,
    row: Dict[str, str],
) -> MachineResult:
    """Full investigation of one machine using Playwright page."""

    result = MachineResult(
        date=row.get("date", ""),
        machine=row.get("machine", ""),
        dai=dai,
        url=row.get("url", ""),
        totalStarts=to_int(row.get("totalStarts")) or 0,
        jackpot=to_int(row.get("jackpot")) or 0,
        continuations=to_int(row.get("continuations")) or 0,
        initialHits=to_int(row.get("initialHits")) or 0,
        maxPayout=to_int(row.get("maxPayout")) or 0,
        existingEstimatedDiffBalls=row.get("estimatedDiffBalls", "").strip(),
    )

    page_url = row.get("url", "")
    if not page_url:
        result.fetchStatus = "url_missing"
        return result

    captured_responses: Dict[str, Any] = {}

    def handle_response(response: Any) -> None:
        url = response.url
        for ep in ["nc-m06-001.php", "nc-m06-003.php", "nc-m06-002.php",
                    "nc-v06-001.php", "nc-v05-001.php"]:
            if ep in url:
                try:
                    ct = response.headers.get("content-type", "")
                    body = response.text()
                    status = response.status
                    try:
                        data = json.loads(body)
                    except json.JSONDecodeError:
                        data = None
                    captured_responses[ep] = {
                        "status": status,
                        "content_type": ct,
                        "body_length": len(body),
                        "is_json": data is not None,
                        "data": data,
                        "body_preview": body[:500] if data is None else None,
                    }
                    save_json(cache_dir / f"{dai}_{ep.replace('.php','')}_response.json",
                              captured_responses[ep])
                except Exception as e:
                    captured_responses[ep] = {"error": str(e)}
                break

    page.on("response", handle_response)

    try:
        response = page.goto(page_url, wait_until="networkidle", timeout=30000)
        if response is None:
            result.fetchStatus = "no_response"
            result.errorType = "navigation_failed"
            return result

        result.httpStatus = str(response.status)

        if response.status in (429, 451, 1015):
            result.fetchStatus = f"blocked_{response.status}"
            result.errorType = "rate_limited"
            return result

        html = page.content()
        save_json(cache_dir / f"{dai}_page_meta.json", {
            "status": response.status,
            "url": response.url,
            "html_length": len(html),
        })

        if any(p in html for p in ["1015", "challenge-platform"]):
            if "challenge-platform" in html and response.status == 200:
                result.notes += "Cloudflare challenge detected; "

        page.wait_for_timeout(3000)

        tokens = extract_tokens_from_html(html)
        ymd_embedded = extract_ymd_biz_from_html(html)
        model = extract_model_from_html(html)

        result.ymdBizEmbedded = ymd_embedded

        if not tokens:
            result.fetchStatus = "token_missing"
            result.errorType = "no_api_tokens"
            with (cache_dir / f"{dai}_page.html").open("w", encoding="utf-8") as f:
                f.write(html)
            return result

        save_json(cache_dir / f"{dai}_tokens.json", {
            "apikey": tokens["apikey"],
            "_i": tokens["_i"][:8] + "...",
            "ymd_embedded": ymd_embedded,
            "model": model,
        })

        page.wait_for_timeout(2000)

        m06_001 = captured_responses.get("nc-m06-001.php")
        if m06_001 and m06_001.get("is_json") and m06_001.get("data"):
            data = m06_001["data"]
            result.m06_001_status = f"ok_json_{m06_001['status']}"
            result.apiResponseKeys = ",".join(sorted(data.keys())) if isinstance(data, dict) else "not_dict"

            diff, pcount, last_raw, gtype, scale = extract_diff_from_m06_001(data)
            result.graphType = gtype
            result.graphPointCount = pcount
            result.graphLastPointRaw = last_raw[:200]
            result.graphScaleInfo = scale

            if diff is not None:
                result.diffBallsFromApi = str(diff)
                result.graphEndpointUsed = "nc-m06-001.php"

            if data.get("redirect_captcha"):
                result.notes += "redirect_captcha=true; "
        elif m06_001:
            result.m06_001_status = f"not_json_{m06_001.get('status','?')}"
            if m06_001.get("body_preview"):
                result.notes += f"m06-001 body preview: {m06_001['body_preview'][:100]}; "

        m06_003 = captured_responses.get("nc-m06-003.php")
        if m06_003 and m06_003.get("is_json") and m06_003.get("data"):
            result.m06_003_status = f"ok_json_{m06_003['status']}"
            diff3, pcount3, last3, scale3 = extract_diff_from_m06_003(m06_003["data"])
            if diff3 is not None and result.diffBallsFromApi == "":
                result.diffBallsFromApi = str(diff3)
                result.graphEndpointUsed = "nc-m06-003.php"
                result.graphPointCount = pcount3
                result.graphLastPointRaw = last3[:200]
                result.graphScaleInfo = scale3
            elif diff3 is not None:
                result.notes += f"m06-003 diff={diff3}; "
        elif m06_003:
            result.m06_003_status = f"not_json_{m06_003.get('status','?')}"

        svg_diff, svg_method, svg_info = extract_svg_diff(page)
        if svg_diff is not None:
            result.diffBallsFromSvg = str(svg_diff)
            result.notes += f"svg: {svg_method}; "
        else:
            result.notes += f"svg: {svg_method}; "

        if result.diffBallsFromApi:
            result.diffBallsFinal = result.diffBallsFromApi
            result.diffBallsMethod = f"api_{result.graphEndpointUsed}"
            result.diffBallsConfidence = "raw_graph_json"
        elif result.diffBallsFromSvg:
            result.diffBallsFinal = result.diffBallsFromSvg
            result.diffBallsMethod = "svg_dom"
            result.diffBallsConfidence = "svg_estimated"
        else:
            result.diffBallsMethod = "none"
            result.diffBallsConfidence = "missing"

        if ymd_embedded and ymd_embedded != target_ymd:
            result.notes += f"YMD_biz mismatch: embedded={ymd_embedded}, target={target_ymd}; "
            if result.diffBallsFinal:
                result.diffBallsConfidence = "wrong_date"
                result.notes += "diff may be for wrong date; "

        result.fetchStatus = "ok" if result.diffBallsFinal else "graph_diff_missing"

    except Exception as e:
        result.fetchStatus = "error"
        result.errorType = type(e).__name__
        result.errorMessage = str(e)[:200]
    finally:
        page.remove_listener("response", handle_response)

    return result


def investigate_from_cache(
    dai: str,
    target_ymd: str,
    cache_dir: Path,
    row: Dict[str, str],
) -> MachineResult:
    """Investigate using only cached data."""
    result = MachineResult(
        date=row.get("date", ""),
        machine=row.get("machine", ""),
        dai=dai,
        url=row.get("url", ""),
        totalStarts=to_int(row.get("totalStarts")) or 0,
        jackpot=to_int(row.get("jackpot")) or 0,
        continuations=to_int(row.get("continuations")) or 0,
        initialHits=to_int(row.get("initialHits")) or 0,
        maxPayout=to_int(row.get("maxPayout")) or 0,
        existingEstimatedDiffBalls=row.get("estimatedDiffBalls", "").strip(),
    )

    m06_001_cache = load_json(cache_dir / f"{dai}_nc-m06-001_response.json")
    if m06_001_cache and m06_001_cache.get("is_json") and m06_001_cache.get("data"):
        data = m06_001_cache["data"]
        result.m06_001_status = f"cached_json_{m06_001_cache.get('status','?')}"
        result.apiResponseKeys = ",".join(sorted(data.keys())) if isinstance(data, dict) else ""
        diff, pcount, last_raw, gtype, scale = extract_diff_from_m06_001(data)
        result.graphType = gtype
        result.graphPointCount = pcount
        result.graphLastPointRaw = last_raw[:200]
        result.graphScaleInfo = scale
        if diff is not None:
            result.diffBallsFromApi = str(diff)
            result.graphEndpointUsed = "nc-m06-001.php"

    m06_003_cache = load_json(cache_dir / f"{dai}_nc-m06-003_response.json")
    if m06_003_cache and m06_003_cache.get("is_json") and m06_003_cache.get("data"):
        result.m06_003_status = f"cached_json_{m06_003_cache.get('status','?')}"
        diff3, pcount3, last3, scale3 = extract_diff_from_m06_003(m06_003_cache["data"])
        if diff3 is not None and result.diffBallsFromApi == "":
            result.diffBallsFromApi = str(diff3)
            result.graphEndpointUsed = "nc-m06-003.php"
            result.graphPointCount = pcount3
            result.graphLastPointRaw = last3[:200]
            result.graphScaleInfo = scale3

    if result.diffBallsFromApi:
        result.diffBallsFinal = result.diffBallsFromApi
        result.diffBallsMethod = f"api_{result.graphEndpointUsed}"
        result.diffBallsConfidence = "raw_graph_json"
    else:
        result.diffBallsMethod = "none"
        result.diffBallsConfidence = "missing"

    result.fetchStatus = "cached_ok" if result.diffBallsFinal else "cached_missing"
    result.notes = "from_cache"
    return result


OUTPUT_COLUMNS = [
    "date", "machine", "dai", "url", "totalStarts", "jackpot", "continuations",
    "initialHits", "maxPayout", "existingEstimatedDiffBalls", "diffBallsFromApi",
    "diffBallsFromSvg", "diffBallsFinal", "diffBallsMethod", "diffBallsConfidence",
    "graphEndpointUsed", "graphPointCount", "graphLastPointRaw", "graphZeroLine",
    "graphScaleInfo", "fetchStatus", "httpStatus", "errorType", "errorMessage",
    "notes", "apiResponseKeys", "graphType", "ymdBizEmbedded",
    "m06_001_status", "m06_003_status",
]


def write_debug_csv(path: Path, results: List[MachineResult]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for r in results:
            writer.writerow({col: getattr(r, col, "") for col in OUTPUT_COLUMNS})


def write_debug_xlsx(path: Path, results: List[MachineResult]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not available, skipping XLSX", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "DiffDebug"
    ws.append(OUTPUT_COLUMNS)
    for r in results:
        ws.append([getattr(r, col, "") for col in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for r_idx in range(2, ws.max_row + 1):
        diff_final = ws.cell(r_idx, OUTPUT_COLUMNS.index("diffBallsFinal") + 1).value
        status = ws.cell(r_idx, OUTPUT_COLUMNS.index("fetchStatus") + 1).value
        if diff_final and str(diff_final).strip():
            for c_idx in range(1, ws.max_column + 1):
                ws.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor="D4EDDA")
        elif status and "blocked" in str(status):
            for c_idx in range(1, ws.max_column + 1):
                ws.cell(r_idx, c_idx).fill = PatternFill("solid", fgColor="F8D7DA")

    wb.save(path)


def print_summary(results: List[MachineResult]) -> None:
    total = len(results)
    api_ok = sum(1 for r in results if r.diffBallsFromApi)
    svg_ok = sum(1 for r in results if r.diffBallsFromSvg)
    final_ok = sum(1 for r in results if r.diffBallsFinal)
    existing_ok = sum(1 for r in results if r.existingEstimatedDiffBalls)
    failed = total - final_ok

    print(f"\n{'='*60}")
    print(f"=== DIFF BALLS INVESTIGATION SUMMARY ===")
    print(f"{'='*60}")
    print(f"Total machines:              {total}")
    print(f"API diff success:            {api_ok}")
    print(f"SVG diff success:            {svg_ok}")
    print(f"Final diff success:          {final_ok}")
    print(f"Failed:                      {failed}")
    print(f"Existing (CSV) diff count:   {existing_ok}")

    status_counts: Dict[str, int] = {}
    for r in results:
        s = r.fetchStatus or "unknown"
        status_counts[s] = status_counts.get(s, 0) + 1
    print(f"\nFetch status breakdown:")
    for s, c in sorted(status_counts.items()):
        print(f"  {s}: {c}")

    error_counts: Dict[str, int] = {}
    for r in results:
        if r.errorType:
            error_counts[r.errorType] = error_counts.get(r.errorType, 0) + 1
    if error_counts:
        print(f"\nError type breakdown:")
        for e, c in sorted(error_counts.items()):
            print(f"  {e}: {c}")

    if existing_ok > 0:
        print(f"\nExisting vs new comparison:")
        for r in results:
            if r.existingEstimatedDiffBalls and r.diffBallsFinal:
                existing = to_int(r.existingEstimatedDiffBalls)
                new = to_int(r.diffBallsFinal)
                if existing is not None and new is not None:
                    delta = abs(new - existing)
                    print(f"  {r.dai}: existing={existing}, new={new}, delta={delta}")

    graph_types: Dict[str, int] = {}
    for r in results:
        gt = r.graphType or "unknown"
        graph_types[gt] = graph_types.get(gt, 0) + 1
    print(f"\nGraph types:")
    for gt, c in sorted(graph_types.items()):
        print(f"  {gt}: {c}")

    ymd_set = set()
    for r in results:
        if r.ymdBizEmbedded:
            ymd_set.add(r.ymdBizEmbedded)
    if ymd_set:
        print(f"\nYMD_biz embedded dates: {sorted(ymd_set)}")


def main() -> int:
    args = parse_args()
    cache_dir = Path(args.cache_dir)
    cache_dir.mkdir(exist_ok=True)

    input_rows = read_csv_input(Path(args.input))
    if args.single:
        input_rows = [r for r in input_rows if str(r.get("dai", "")).zfill(4) == args.single.zfill(4)]

    results: List[MachineResult] = []

    if args.skip_browser:
        print(f"Cache-only analysis of {len(input_rows)} machines...")
        for row in input_rows:
            dai = str(row.get("dai", "")).zfill(4)
            r = investigate_from_cache(dai, args.date, cache_dir, row)
            results.append(r)
            status_icon = "O" if r.diffBallsFinal else "X"
            print(f"  [{status_icon}] {dai}: {r.fetchStatus} diff={r.diffBallsFinal or 'N/A'}")
    else:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=args.headless)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36",
                viewport={"width": 412, "height": 915},
                locale="ja-JP",
            )
            page = context.new_page()

            processed = 0
            stopped = False

            for row in input_rows:
                if args.limit and processed >= args.limit:
                    break
                if stopped:
                    break

                dai = str(row.get("dai", "")).zfill(4)

                if args.resume:
                    cached = load_json(cache_dir / f"{dai}_nc-m06-001_response.json")
                    if cached and cached.get("is_json"):
                        r = investigate_from_cache(dai, args.date, cache_dir, row)
                        results.append(r)
                        status_icon = "O" if r.diffBallsFinal else "X"
                        print(f"  [CACHE {status_icon}] {dai}: {r.fetchStatus}")
                        continue

                print(f"  Investigating {dai}...", end=" ", flush=True)
                r = investigate_machine_with_browser(page, dai, args.date, cache_dir, row)
                results.append(r)
                processed += 1

                status_icon = "O" if r.diffBallsFinal else "X"
                print(f"[{status_icon}] {r.fetchStatus} http={r.httpStatus} diff={r.diffBallsFinal or 'N/A'} gtype={r.graphType}")

                if r.fetchStatus and "blocked" in r.fetchStatus:
                    print(f"  BLOCKED! Stopping. Reason: {r.fetchStatus}")
                    stopped = True
                    break

                if processed < len(input_rows):
                    time.sleep(args.delay)

            browser.close()

    write_debug_csv(Path(args.out_csv), results)
    write_debug_xlsx(Path(args.out_xlsx), results)
    print_summary(results)

    summary = {
        "total": len(results),
        "api_diff_ok": sum(1 for r in results if r.diffBallsFromApi),
        "svg_diff_ok": sum(1 for r in results if r.diffBallsFromSvg),
        "final_diff_ok": sum(1 for r in results if r.diffBallsFinal),
        "existing_diff_ok": sum(1 for r in results if r.existingEstimatedDiffBalls),
        "failed": sum(1 for r in results if not r.diffBallsFinal),
        "status_counts": {},
        "machines": [asdict(r) for r in results],
    }
    for r in results:
        s = r.fetchStatus or "unknown"
        summary["status_counts"][s] = summary["status_counts"].get(s, 0) + 1
    save_json(Path(args.out_csv.replace(".csv", "_summary.json")), summary)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

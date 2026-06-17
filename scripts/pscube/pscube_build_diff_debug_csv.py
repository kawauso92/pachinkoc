#!/usr/bin/env python3
"""
Build the diff-ball debug CSV from all existing data sources.
No network requests - purely offline analysis.
"""
import csv
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional


OUTPUT_COLUMNS = [
    "date", "machine", "dai", "url", "totalStarts", "jackpot", "continuations",
    "initialHits", "maxPayout", "existingEstimatedDiffBalls",
    "diffBallsFromApi", "diffBallsFromSvg", "diffBallsFinal",
    "diffBallsMethod", "diffBallsConfidence",
    "graphEndpointUsed", "graphPointCount", "graphLastPointRaw",
    "graphZeroLine", "graphScaleInfo",
    "fetchStatus", "httpStatus", "errorType", "errorMessage", "notes",
]


def to_int(v: Any) -> Optional[int]:
    if v is None: return None
    s = str(v).strip().replace(",", "")
    if not s: return None
    try: return int(float(s))
    except ValueError: return None


def load_json(path: Path) -> Any:
    if not path.exists(): return None
    with path.open("r", encoding="utf-8-sig") as f:
        return json.load(f)


def main():
    input_csv = Path("data/pscube/samples/arrow_mihara_eva_4p_20260616.csv")
    history_cache = load_json(Path("data/pscube/samples/pscube_history_cache_20260616.json"))
    diff_probe = load_json(Path("data/pscube/samples/pscube_diff_cache_probe_20260616.json"))

    # Build history lookup
    history_by_dai = {}
    if history_cache and isinstance(history_cache, dict):
        for m in history_cache.get("machines", []):
            dai = str(m.get("dai", "")).zfill(4)
            history_by_dai[dai] = m

    # Build diff probe lookup
    diff_by_dai = {}
    if diff_probe and isinstance(diff_probe, dict):
        for m in diff_probe.get("machines", []):
            dai = str(m.get("dai", "")).zfill(4)
            diff_by_dai[dai] = m

    # Build debug cache lookup
    debug_cache = Path("debug_cache")
    debug_by_dai = {}
    if debug_cache.exists():
        for f in debug_cache.glob("*_nc-m06-001_response.json"):
            dai = f.name.split("_")[0]
            debug_by_dai[dai] = load_json(f)

    # Read input CSV
    with input_csv.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))

    results = []
    for row in rows:
        dai = str(row.get("dai", "")).zfill(4)
        existing_diff = row.get("estimatedDiffBalls", "").strip()

        r = {
            "date": row.get("date", ""),
            "machine": row.get("machine", ""),
            "dai": dai,
            "url": row.get("url", ""),
            "totalStarts": row.get("totalStarts", ""),
            "jackpot": row.get("jackpot", ""),
            "continuations": row.get("continuations", ""),
            "initialHits": row.get("initialHits", ""),
            "maxPayout": row.get("maxPayout", ""),
            "existingEstimatedDiffBalls": existing_diff,
            "diffBallsFromApi": "",
            "diffBallsFromSvg": "",
            "diffBallsFinal": "",
            "diffBallsMethod": "",
            "diffBallsConfidence": "",
            "graphEndpointUsed": "",
            "graphPointCount": "",
            "graphLastPointRaw": "",
            "graphZeroLine": "",
            "graphScaleInfo": "",
            "fetchStatus": "",
            "httpStatus": "",
            "errorType": "",
            "errorMessage": "",
            "notes": "",
        }

        notes = []

        # Check history cache
        hc = history_by_dai.get(dai)
        if hc:
            h_status = hc.get("fetchStatus", "")
            h_update = hc.get("updateText", "")
            h_hist = len(hc.get("historyRows", []))
            notes.append(f"history_cache: status={h_status}, histRows={h_hist}, update={h_update}")

            # Check if data table has relevant info
            for dr in hc.get("dataRows", []):
                if isinstance(dr, list) and dr and dr[0] == "累計スタート":
                    starts_today = dr[1] if len(dr) > 1 else ""
                    starts_1d = dr[2] if len(dr) > 2 else ""
                    notes.append(f"starts: today={starts_today}, 1d_ago={starts_1d}")
                elif isinstance(dr, list) and dr and dr[0] == "最大放出数":
                    payout_today = dr[1] if len(dr) > 1 else ""
                    payout_1d = dr[2] if len(dr) > 2 else ""
                    notes.append(f"maxPayout: today={payout_today}, 1d_ago={payout_1d}")

        # Check diff probe cache
        dp = diff_by_dai.get(dai)
        if dp:
            dp_status = dp.get("fetchStatus", "")
            notes.append(f"diff_probe: {dp_status}")
            r["fetchStatus"] = dp_status
            if dp_status.startswith("page_http_"):
                r["httpStatus"] = dp_status.replace("page_http_", "")

        # Check debug cache
        dc = debug_by_dai.get(dai)
        if dc:
            dc_status = dc.get("status", "")
            dc_is_json = dc.get("is_json", False)
            notes.append(f"debug_cache: status={dc_status}, is_json={dc_is_json}")
            if dc_status:
                r["httpStatus"] = str(dc_status)
            if not dc_is_json:
                r["errorType"] = "api_not_json"
                r["errorMessage"] = (dc.get("body_preview", "") or "")[:100]
                r["fetchStatus"] = f"api_blocked_{dc_status}"

        # Determine final diff
        if existing_diff:
            r["diffBallsFinal"] = existing_diff
            r["diffBallsMethod"] = "csv_existing"
            r["diffBallsConfidence"] = "csv_manual"
            r["fetchStatus"] = r["fetchStatus"] or "existing_in_csv"
            notes.append("diff from original CSV (likely manual entry or codex browser)")
        else:
            r["fetchStatus"] = r["fetchStatus"] or "no_diff_source"
            r["diffBallsMethod"] = "none"
            r["diffBallsConfidence"] = "missing"

            # Determine specific failure reason
            if dc and dc.get("status") == 451:
                r["errorType"] = "api_451_blocked"
                r["errorMessage"] = "nc-m06-001.php returned 451 (WAF/bot block)"
            elif dp and dp.get("fetchStatus", "").startswith("page_http_502"):
                r["errorType"] = "page_502"
                r["errorMessage"] = "Machine page returned 502 Bad Gateway"
            elif not dc and not dp:
                r["errorType"] = "never_attempted"
                r["errorMessage"] = "No fetch attempt recorded for this machine"

        r["notes"] = " | ".join(notes)
        results.append(r)

    # Write CSV
    out_csv = Path("data/pscube/samples/pscube_mihara_eva_diff_debug_20260616.csv")
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for r in results:
            writer.writerow(r)

    # Write XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "DiffDebug"
        ws.append(OUTPUT_COLUMNS)
        for r in results:
            ws.append([r.get(c, "") for c in OUTPUT_COLUMNS])

        hdr = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.fill = hdr
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for ri in range(2, ws.max_row + 1):
            diff_col = OUTPUT_COLUMNS.index("diffBallsFinal") + 1
            diff_val = ws.cell(ri, diff_col).value
            if diff_val and str(diff_val).strip():
                for ci in range(1, ws.max_column + 1):
                    ws.cell(ri, ci).fill = PatternFill("solid", fgColor="D4EDDA")
            else:
                for ci in range(1, ws.max_column + 1):
                    ws.cell(ri, ci).fill = PatternFill("solid", fgColor="F8D7DA")

        wb.save("data/pscube/samples/pscube_mihara_eva_diff_debug_20260616.xlsx")
        print("XLSX saved.")
    except ImportError:
        print("openpyxl not available, XLSX skipped.")

    # Summary
    total = len(results)
    has_diff = sum(1 for r in results if r["diffBallsFinal"])
    no_diff = total - has_diff

    error_types = {}
    for r in results:
        et = r["errorType"] or ("ok" if r["diffBallsFinal"] else "unknown")
        error_types[et] = error_types.get(et, 0) + 1

    print(f"\n{'='*60}")
    print(f"DIFF BALLS DEBUG CSV SUMMARY")
    print(f"{'='*60}")
    print(f"Total machines: {total}")
    print(f"Has diff (from CSV): {has_diff}")
    print(f"Missing diff: {no_diff}")
    print(f"\nError breakdown:")
    for et, cnt in sorted(error_types.items(), key=lambda x: -x[1]):
        print(f"  {et}: {cnt}")

    # Machine list
    print(f"\n{'='*60}")
    print(f"Machine-by-machine status:")
    print(f"{'='*60}")
    for r in results:
        icon = "O" if r["diffBallsFinal"] else "X"
        diff = r["diffBallsFinal"] or "N/A"
        err = r["errorType"] or ""
        print(f"  [{icon}] {r['dai']} {r['machine'][:12]:12s} diff={diff:>8s}  {err}")

    print(f"\nCSV: {out_csv}")


if __name__ == "__main__":
    main()
